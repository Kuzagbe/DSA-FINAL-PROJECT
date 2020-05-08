from openpyxl import *

#THIS IS A PROGRAM THAT HELP COMPANY TO RENT CARS

#Below I have difined all dictionalies that are need to hold detaiels of all cars
###############################CAR############################################
#Below dictionaly holds all modeles of cars
car_model = {'car1': 'Benz-2xs', 'car2': 'Audi-9f', 'car3': 'Dodge-Ram','car4': 'Mazda-12', 'car5': 'Polo-cup'}
#Below dictionaly holds all released years of cars
year_released = {'car1': '2012', 'car2': '2013', 'car3': '2015', 'car4': '2010', 'car5': '2016'}
#Below dictionaly holds all acquired years of cars
year_acquired = {'car1': '2014', 'car2': '2015', 'car3': '2017', 'car4': '2013', 'car5': '2018'}
#Below dictionaly holds all money made from cars
money_made = {'car1': 0, 'car2': 0, 'car3': 0, 'car4': 0, 'car5': 0}
#Below dictionaly holds all plate numbers of cars
plate_number = {'car1': 'RAC900U', 'car2': 'RAD100Y', 'car3': 'RAD565Y', 'car4': 'RAD54E', 'car5': 'RAD454I'}
#Below dictionaly holds all renting times of cars
rent_number = {'car1': 0, 'car2': 0, 'car3': 0, 'car4': 0, 'car5': 0}
#Below dictionaly holds stutus of cars either AVILABEL OR ON RENT
status = {'car1': 'AVAILABLE', 'car2': 'AVAILABLE', 'car3': 'AVAILABLE', 'car4': 'AVAILABLE', 'car5': 'AVAILABLE'}
#Below dictionaly holds fixed prices/day of cars
prices={'car1': 100, 'car2': 300, 'car3': 200, 'car4': 150, 'car5': 400}

###############################CUSTOMER############################################

cst_name = {'cst1': 'Gilbert', 'cst2': 'Cofie', 'cst3':'Alan'}
#Below dictionaly holds all released years of cars
cust_address = {'cst1': 'Remera-kigali', 'cst2': 'Accra-Ghana', 'cst3': 'Kampala-Uganda'}
#Below dictionaly holds all acquired years of cars
cst_contact = {'cst1': '8787879', 'cst2': '78787879', 'cst3': '788789898'}
#Below dictionaly holds all money made from cars
Cust_Car_model = {'cst1':'Benz 2xs', 'cst2':'Mazda-12', 'cst3':'Audi 9f'}
#Below dictionaly holds all plate numbers of cars
Renting_days = {'cst1': 3, 'cst2': 2, 'cst3': 2}



#this function helps to exit the program
def exit_this():
    exit("THANK YOU FOR USING THIS SOFTWARE SEE YOUUU!!!")
#this function helps to add a new car
def add_car():
    carn = 'car' + str(len(car_model) + 1) #this helps to create new key for new car that is going to be added
    car_mod = input("Enter Car model: ") #helps to add new modle of new car
    year_rlzed = input("Enter a year that Car has been released in: ") #add year realesed of new car
    year_acqrd = input("Enter a year that Car has been acquired in: ")#add year acquired of new car
    plate_num = input("Enter a plate number of that Car: ")  #add plate number of new car
    price = input("Enter fixed price/day of this car for rent: ") #add price of new car

    car_model[carn] = car_mod #assign model to new created key
    year_released[carn] = year_rlzed #assign year released to new created key
    year_acquired[carn] = year_acqrd #assign year acquired to new created key
    plate_number[carn] = plate_num #assign plate number to  a new created key
    status[carn] = 'AVAILABLE' #assign status to new created key
    prices[carn] = int(price) #assign price to new created key
    money_made[carn]=0 #assign 0 money to new created key
    rent_number[carn]=0 #assign 0 renting times to a new created key

    wb = load_workbook("Reportog.xlsx")
    ws = wb.active
    ws['A7'] = car_mod
    ws['B7'] = 0
    ws['C7'] = 0

    wb.save("Reportog.xlsx")



#this function helps to rent a car
def rent_car():
    index = 1
    for x in car_model:  #for loop to read all car model
        # display of car model and status
        print(index, ':', car_model[x],'--',status[x],'--',prices[x])
        index += 1
        # prompting user to select a car he/she want to rent
    rent = eval(input("Select car by number assigned to them(ex:1): "))
    carn = 'car' + str(rent)
    for i in car_model: #loop to check availability and allow selection to go smothly
        # checking if car is available
        if carn.casefold() == i.casefold() and status[i] == 'AVAILABLE':
            #asking user to choose b2n fixed/negotiable price
            days=eval(input('Enter number of days you want to rent this car'))
            # checking if it is negotiable price
            #chacking if it is fixed prices
            # adding one time to renting times
            rent_number[i] = rent_number[i] + 1
            # adding money
            realprice=days*prices[i]
            money_tmp = realprice + money_made[i]
            money_made[i] = money_tmp
            # changing status to 'ON RENT'
            status[i] = 'ON RENT'
            # Dsplaying updates on car
            print('For now ', car_model[i], ' has been rented on ', money_made[i], ' in total')
            print('And it has been rented', rent_number[i], 'times')
            wb = load_workbook("Reportog.xlsx")
            ws = wb.active
            for row in ws.iter_rows(2):
                for cell in row:
                    if cell.value == car_model[i]:
                        ws.cell(row=cell.row, column=2).value = ws.cell(row=cell.row, column=2).value + realprice
                        ws.cell(row=cell.row, column=3).value = ws.cell(row=cell.row, column=3).value + 1
                        wb.save("Reportog.xlsx")

            print("___________________")
            print('| CUSTOMER DETAILS ')  # tabs and decorations
            print("|------------------")
            cstn = 'cst' + str(len(cst_name) + 1)  # this helps to create new key for new car that is going to be added
            cst_na = input("Enter Customer name: ")  # add customer names
            cust_add = input("Enter Customer addresses:")  # add customer age
            cst_cont = input("Enter Customer contact: ")  # add customer address
            Cust_model = car_model[i]
            Renting = days

            cst_name[cstn] = cst_na  # assign model to new created key
            cust_address[cstn] = cust_add  # assign year released to new created key
            cst_contact[cstn] = cst_cont # assign year acquired to new created key
            Cust_Car_model[cstn] = Cust_model   # assign plate number to  a new created key
            Renting_days[cstn] = Renting  # assign status to new created key



               #chacking if user choose a car that is on rent
        elif carn.casefold() == i.casefold() and status[i]=='ON RENT':

            print('This car is on rent now!!!')

#this function helps to remove a car
def remov_car():
    #text to remind user to selecte a car he/she wants to remove
    print('Select a car you want to remove')
    index=1
    for v in car_model: #loop to display all cars
        print(index, ':', car_model[v]) #print all cars
        index += 1
        #prompting user to select car he/she wants to remove
    remov = eval(input("Select you want to remove(ex:1): "))
    carn = 'car' + str(remov)
    for f in list(car_model.keys()): #loop to delete cars, dectionsly was conerted into list to delete
        if carn.casefold() == f.casefold():  #checking if selected car belongs to dictionaly carmodel
            del car_model[f]      #deleting car model
            del year_released[f]  #deleting  year released of car
            del year_acquired[f]  #deleting  year released of car
            del plate_number[f]   #deleting  plate number of car
            del rent_number[f]    #deleting  renting times of car
            del money_made[f]     #deleting  money made from car
            del prices[f]         #deleting  price of car
            del status[f]         #deleting  price of car
            print("YOUR CAR HAS SUCCESSFULLY REMOVED FROM THE LIST!!")

#this function helps to check all cars their renting times, and money made from them
def check_trans():
    print("___________________________________")
    print('| Car          Money     Rent times ')  #tabs and decorations
    print("|-----------------------------------")
    for j in car_model:                           #loop to desplay all cars through car model
        #display car model, money made from them and renting times
        print('|', car_model[j], '--->', money_made[j], '--->', rent_number[j])

def check_Cutomer():
    print("_________________________________________________________________________")
    print('| Name          Contacts            Address        Car-model       Days ')  # tabs and decorations
    print("|------------------------------------------------------------------------")
    for i in cst_name:  # loop to display all cars through customer names
        # Display customer details
        print('|', cst_name[i], '--->', cst_contact[i], '--->', cust_address[i], '--->', Cust_Car_model[i], '---->',
              Renting_days[i])


    print("-----------------------------------") #decoration
#this function display all cars and main details of them
def all_cars():
    #decorations
    print("_______________________________________________________________________")
    print('| Car         Plate_number Released  Acquired  Money   Rent times')
    print("|----------------------------------------------------------------------")
    for op in car_model: #loops to display all cars through car model
        #display all cars and main details of them
        print('|', car_model[op], '--', plate_number[op], '--', year_released[op], '--',
              year_acquired[op], '--', money_made[op], '--', rent_number[op])


    print("|----------------------------------------------------------------------")
#this function helps to put cars on rent, back on the cue
def put_on_cue():
    index = 1
    for i in status:   #loops to display status
        if status[i]=='ON RENT':   #condition to check if car has ON RENT status
            print(index,':',car_model[i],'--',status[i])    #display cars on rent
            index += 1
            #prompting user to select car he/sh wants to put back on cue
    put_back=eval(input('Select you want to put back on cue(ex:1) :'))
    car = 'car' + str(put_back)
    for x in status: #loop to find car user spacified
        if car.casefold() == x.casefold():   #check if car user spaified is on the list
            status[x] = 'AVAILABLE'  #changing status of car
            print("YOUR CAR HAS SUCCESSFULLY ADDED BACK ON CUE!!") #Notifying user




# this function helps to display menu
def main():
    print("USE NUMBER TO SELECT ANY THING FROM MENU")
    print("____________________________________________")
    print("|\t 1.ADD NEW CAR                          |")
    print("|\t 2.RENT A CAR                           |")
    print("|\t 3.REMOVE A CAR                         |")
    print("|\t 4.CHECK NUMBER OF TIMES A CAR HAS      |")
    print("|\t   BEEN RENTED, AND MONEY MADE FROM IT  |")
    print("|\t 5.DISPLAY ALL CARS IN YOUR STORE       |")
    print("|\t 6.PUT A CAR BACK ON CUE                |")
    print("|\t 7.DISPLAY CUSTOMER DETAILS             |")
    print("|___________________________________________|")
     #prompting user to select one of the menu
    selection = input("Select menu with using number(1-6): ")

    if selection == '1': #checking if user selected one
        add_car()        #calling function to add car
    elif selection == '2':  #checking if user selected two
        rent_car()          #calling function to rent car
    elif selection == '3':  #checking if user selected three
        remov_car()         #calling function to remove car
    elif selection == '4':  #checking if user selected four
        check_trans()       #calling function to check cars, money made and renting times
    elif selection == '5':  #checking if user selected five
        all_cars()          #calling function to display cars
    elif selection == '6':  #checking if user selected six
        put_on_cue()        #calling function to put car back on cue
    elif selection == '7':  #checking if user selected seven
        check_Cutomer()        #calling function to Check cusomer details
    else:
        #if user doesn't  choose any of above numbers
        print("NO CHOICE WE HAVE THAT LOOK LIKE THAT")

#calling function to display menu
main()
y = True #asigning true to y
while y == y:  #creating endless loop
    #prompting user to select Yes to continue or No to exit
    y = input("If you want to CONTINUE enter Y and if  you want to EXIT enter N : ")
    if y.casefold() == "Y".casefold(): #checking if user choose yes
        main()                          #calling main function
    elif y.casefold() == 'N'.casefold(): #checking if user choose no
        exit_this()                      #calling function to teriminate program
    else:
        #text to display if user doesn't choose any
        print("You did not choose any ")
        exit()          #clos program
